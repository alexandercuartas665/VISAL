"""
Build-ConsentimientosExcel.py

Para cada .pdf en el directorio CONSENTIMIENTOS:
  - Crea una subcarpeta con el mismo nombre base del archivo
  - Extrae cada pagina como PNG con PyMuPDF (fitz)

Luego genera un Excel con las columnas:
  Codigo | Archivo | Ruta | Nombre en Sistema | ID Sistema | Carpeta capturas |
  Total paginas | Prompt de construccion

El prompt de construccion se autogenera con el patron que estamos usando
para reconstruir cada consentimiento (con placeholders para ajustes).
"""

import fitz  # PyMuPDF
import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------- Config ----------------
DIR_BASE = r"C:\Users\acuartas\OneDrive - Bitcode IT Services S.A.S\Bitcode\13. Proyectos\028. Visal\05. Formatos\CONSENTIMIESNTOS"
XLSX_OUT = os.path.join(DIR_BASE, "_INVENTARIO_CONSENTIMIENTOS.xlsx")
DPI = 150  # resolucion de las paginas exportadas

# ---------------- Consentimientos del sistema ----------------
# Volcado directo desde la BD (codigo, nombre, id)
SISTEMA = [
    ("PP-FO-112", "PP-FO-112 FORMATO CONSENTIMIENTO MEDICINA FÍSICA Y REHABILITACIÓN", "ea39e643-38a4-487f-a866-aa40c4c4a8c5"),
    ("PP-FO-113", "PP-FO-113 FORMATO CONSENTIMIENTO INFORMADO PARA CONSULTA DE PRIMERA VEZ", "2679dac5-a941-4246-aaab-7efcf19db660"),
    ("PP-FO-17",  "PP-FO-17 FORMATO CONSENTIMIENTO ENFERMERIA", "37b80bcd-5ebb-4b3c-8ab5-9fbd7bd151d4"),
    ("PP-FO-18",  "PP-FO-18 FORMATO CONSENTIMIENTO TERAPIAS", "ec989dd1-1294-4ac3-b90c-44837f1ecebc"),
    ("PP-FO-20",  "PP-FO-20 FORMATO CONSENTIMIENTO MED GRAL", "a75d37e4-c7a3-453b-94d7-25b2b5ad181c"),
    ("PP-FO-22",  "PP-FO-22 FORMATO CONSENTIMIENTO PSICOLOGIA", "db2fa994-0aa1-4f16-8910-bc30ac7ec078"),
    ("PP-FO-23",  "PP-FO-23 FORMATO DE DESISTIMIENTO DE SERVICIOS VISAL RT", "6c8670f1-bab2-4f80-b02e-6d52b025dda1"),
    ("PP-FO-24",  "PP-FO-24 FORMATO CONSENTIMIENTO COVID 19", "24390d26-31df-4d65-9775-328cef43bce2"),
    ("PP-FO-32",  "PP-FO-32 CONSENTIMIENTO INFORMADO MEDICINA LABORAL", "ba747413-0a72-4371-a1c9-53dd361b4a3a"),
    ("PP-FO-35",  "PP-FO-35 FORMATO INGRESO PROGRAMA DE ATENCIÓN DOMICILIARIAV3", "91ecfd4b-c460-45a1-a36a-26b35437b17f"),
    ("PP-FO-37",  "PP-FO-37 FORMATO DE DESESCALONAMIENTO EGRESO PAD", "b0307987-7576-42ea-8713-023411af2b19"),
    ("PP-FO-37-PAD", "PP-FO-37 FORMATO DE DESESCALONAMIENTO PAD", "d95acba3-9385-4053-93ba-b639c6401391"),
    ("PP-FO-66",  "PP-FO-66 FORMATO NO REANIMACION", "1854f5ef-98c5-4a89-a13e-7265fb01d357"),
    ("PP-FO-69",  "PP-FO-69 FORMATO CONSENTIMIENTO PRUEBA DIAGNOSTICA DE VIH", "f6550d6e-435d-46da-90c4-a0f6d4623ccf"),
    ("PP-FO-81",  "PP-FO-81 FORMATO CONSENTIMIENTO TRABAJO SOCIAL", "3115886e-76f2-463d-93bb-82331b00234b"),
    ("PP-FO-88",  "PP-FO-88 CONSENTIMIENTO INFORMADO NUTRICION", "c5c3e0f1-730e-4035-9ff7-6524952b962b"),
    ("PP-FO-89",  "PP-FO-89 FORMATO CONSENTIM MED INTERNA", "031d4e9d-c314-4389-b6e7-51b7bb8c47e9"),
    ("PP-FO-90",  "PP-FO-90 FORMATO CONSENTIM MED  FAMILIAR", "c1c8690a-19d5-4bfc-a7b3-9c8b2220e790"),
    ("PP-FO-96",  "PP-FO-96 FORMATO CONSENTIMIENTO ORTOPEDIA Y TRAUMATOLOGÍA MODALIDAD TELEMEDICINA.", "40c7f480-a032-4f56-9360-5dd1b21eb684"),
]

def normaliza_codigo(nombre_archivo: str) -> str | None:
    """Extrae 'PP-FO-XXX' del inicio del nombre del archivo."""
    m = re.match(r"^(PP-FO-\d+[a-zA-Z\-]*)\b", nombre_archivo)
    if not m:
        return None
    codigo = m.group(1)
    # Los dos PP-FO-37 se diferencian por si el nombre contiene "PAD" y no "EGRESO"
    if codigo == "PP-FO-37":
        if "PAD" in nombre_archivo.upper() and "EGRESO" not in nombre_archivo.upper():
            return "PP-FO-37-PAD"
    return codigo


def buscar_en_sistema(codigo: str) -> tuple[str, str] | None:
    """Retorna (nombre, id) del consentimiento en el sistema o None."""
    for cod, nom, sid in SISTEMA:
        if cod == codigo:
            return (nom, sid)
    return None


def prompt_construccion(codigo: str, nombre_archivo: str, subfolder: str, npag: int) -> str:
    """Autogenera el prompt para reconstruir el consentimiento a mano."""
    return f"""Reconstruye el consentimiento {codigo} en el motor de formularios de Visal (form_definitions, tipo=CONSENTIMIENTO) siguiendo estas pautas:

## Contexto
- El schema en jsonb debe replicar EXACTAMENTE la estructura del documento Word original:
    {nombre_archivo}
- Referencia visual: cada pagina del docx esta capturada como PNG en la carpeta:
    {subfolder}
  ({npag} pagina(s) exportadas). Miralas ANTES de escribir cualquier codigo — te ayudan a distinguir tablas reales de campos sueltos.

## Reglas OBLIGATORIAS de interpretacion del docx

### R1. Header y firma profesional — NO SE TOCAN
- La seccion de datos del paciente (header) YA esta correcta en el sistema como "Datos del Paciente (auto-llenado)" (id="auto-datos-paciente"). NO la reescribas, NO la muevas, NO cambies sus 5 campos:
    nombre_paciente_consent (text), tipo_documento_consent (text),
    numero_documento_consent (text), edad_consent (number), fecha_atencion_consent (date).
- La seccion de firma profesional YA esta correcta como "Firmas (auto-llenadas)" con firma_paciente_consent y firma_profesional_consent. NO la reescribas.
- PERO respeta el NUMERAL/orden del docx: si en el docx la firma profesional es el punto 5 o 6, ubicala como ultima seccion (antes de "Cierre" y "MEDICO" que van al final por convencion del sistema).

### R2. Tablas con ultima columna VACIA sin titulo = casilla para marcar con X
Cuando veas una tabla del docx con una ultima columna en blanco (sin encabezado), es una casilla para marcar. NO la ignores. Modelala asi:
  - Agrega una columna extra al final llamada "Marca" con name="marca", fieldType="text", defaultValue=""
  - Ejemplo tipico: tabla PROCEDIMIENTOS con columnas [Procedimiento, Descripcion, <vacia>] => 3 columnas donde la 3era es "Marca" (text libre para X o vacio)
  - Ejemplo tipico: tabla BENEFICIOS con [Beneficio, <vacia>] => 2 columnas donde la 2da es "Marca"

### R3. Guiones bajos (____________) = campo de entrada del usuario
Cualquier campo del docx con formato "Etiqueta: ______________" o "Etiqueta ___________" es una entrada que el operador va a llenar en el sistema. NO lo dejes como texto plano dentro de un parrafo. Conviertelo a un field:
  - Etiqueta con espacio corto => field/text con widthColumns segun tamano visual
  - Textos largos multilinea (parrafos con muchos guiones) => field/textarea con enableVoice=true
  - Fechas => field/date
  - "Nombre: ________, CC: ______" => 2 fields separados en linea
  - "SI___ O NO___" o "Si o No" => field/select con options=["Si","No"]

### R4. Tablas del docx = tablas del schema (nunca campos sueltos)
Cada tabla visual del docx se convierte en fieldType="table" con seedRows locked (lockRows=true), NUNCA en campos individuales sueltos. Si ves una tabla con celdas etiquetadas (Item / Observacion / etc), es una tabla seed. Confirma con las capturas de {subfolder}.

### R5. Datos prellenados "NO REFIERE" u otros defaults
Cuando en el docx aparece "NO REFIERE" o valores similares como default en la columna Observacion, usala como:
    fieldType="select"
    options=["NO REFIERE","REFIERE"] (u opciones apropiadas segun contexto)
    allowCustom=true
    defaultValue="NO REFIERE"
Asi el operador ve la lista pero puede escribir libre encima.

### R6. HUELLA dactilar
Si el docx tiene cajas rectangulares para huella dactilar, usa `field/textarea` con label="HUELLA", widthColumns=3, rows=4, placeholder="Espacio para huella". NUNCA texto plano "HUELLAHUELLA" en un parrafo.

### R7. Textareas evaluables
Para conceptos, observaciones, planes, analisis y textos largos usa fieldType="textarea" con enableVoice=true (dictado por voz) para todos los formularios del sistema.

### R8. Declaraciones del paciente / responsable / profesional
Preserva las declaraciones textuales COMO ESTAN en el docx (mismo texto, respetando comas, saltos de linea). Solo elimina prefijos residuales tipo "HUELLAHUELLA" que pudieran quedar del parser.

### R9. Numerales del docx en subheadings
Si el docx numera secciones (1. DATOS, 2. INFORMACION SOBRE EL PROCEDIMIENTO, 3. DECLARACION DEL PACIENTE...), respetar el numero en los subheadings del schema (content="2. INFORMACION SOBRE EL PROCEDIMIENTO"). Ayuda al operador a ubicarse.

## Como escribir el rework
Sigue el patron de los scripts ya probados en:
    C:\\DesarrolloIA\\Visal\\scripts\\Rework-HCFO*-*.ps1 y Add-PPFO113-Consentimiento.ps1
Un solo script PowerShell que:
    - carga el schema actual (docker exec psql)
    - **preserva las secciones "Datos del Paciente (auto-llenado)" y "Firmas (auto-llenadas)" si existen** (o las agrega si el consentimiento no las tiene aun)
    - construye el schema nuevo con hashtables
    - hace UPDATE sobre form_definitions.schema_json via docker exec psql
    - imprime el diff resultante

Guarda backup previo en C:\\Users\\acuartas\\AppData\\Local\\Temp\\* antes del UPDATE.

**IMPORTANTE — bug conocido de PowerShell**: al construir seedRows con UNA sola fila, PowerShell aplana el eje externo. Usa el helper `Tabla` con `[System.Collections.ArrayList]` como en los scripts existentes, o corrige con UPDATE SQL posterior (mira patrones en fix14.sql / fix19.sql en scratchpad).

## Verificacion
1. HC-FO-08 debe quedar intacto (diff verificado).
2. Abrir /formularios en localhost:5080 y confirmar que el consentimiento carga sin errores (usar MCP Chrome).
3. Verificar que "TABLA REPETIBLE" aparece por cada tabla seed esperada.
4. Verificar que la seccion "Datos del Paciente (auto-llenado)" aparece PRIMERA y con los 5 campos.
5. Verificar que las columnas de MARCA X estan presentes en las tablas donde el docx muestra la ultima columna vacia.
6. Verificar que los guiones bajos se convirtieron en fields de entrada (no quedaron como texto).
7. Verificar que "Firmas (auto-llenadas)" esta al final antes de Cierre y MEDICO.

## Al terminar
Commit con: feat(consentimientos): rework {codigo} con tablas seed, marca X y campos de entrada
""".strip()


# ---------------- Proceso ----------------
def main():
    dir_base = Path(DIR_BASE)
    if not dir_base.exists():
        print(f"ERROR: no existe {DIR_BASE}")
        return

    pdfs = sorted(dir_base.glob("*.pdf"))
    print(f"Encontrados {len(pdfs)} PDFs")

    filas = []
    for pdf in pdfs:
        # Trim el nombre para evitar problemas con carpetas en Windows
        base = pdf.stem.strip().rstrip(".").strip()
        subfolder = dir_base / base
        subfolder.mkdir(exist_ok=True)

        # Renderizar cada pagina
        doc = fitz.open(pdf)
        npag = len(doc)
        for i, page in enumerate(doc, start=1):
            pngpath = subfolder / f"pagina_{i:02d}.png"
            if not pngpath.exists():
                pix = page.get_pixmap(dpi=DPI)
                pix.save(str(pngpath))
        doc.close()
        print(f"  {base}: {npag} pagina(s) OK")

        # Match con sistema
        codigo = normaliza_codigo(pdf.name)
        info = buscar_en_sistema(codigo) if codigo else None
        nombre_sis, id_sis = (info if info else ("NO ENCONTRADO", ""))

        filas.append({
            "codigo": codigo or "?",
            "archivo": pdf.name.replace(".pdf", ".docx"),
            "ruta_docx": str(dir_base / pdf.name.replace(".pdf", ".docx")),
            "nombre_sistema": nombre_sis,
            "id_sistema": id_sis,
            "carpeta_capturas": str(subfolder),
            "npag": npag,
            "prompt": prompt_construccion(codigo or "?", pdf.name.replace(".pdf",".docx"), str(subfolder), npag),
        })

    # -------- Excel --------
    wb = Workbook()
    ws = wb.active
    ws.title = "Consentimientos"

    headers = [
        "Codigo Sistema",
        "Archivo Word",
        "Ruta Word",
        "Nombre en Sistema",
        "ID Sistema (UUID)",
        "Carpeta de Capturas",
        "Total Paginas",
        "Prompt de Construccion",
    ]
    thick = Side(border_style="thin", color="666666")
    bord = Border(left=thick, right=thick, top=thick, bottom=thick)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bord
    ws.row_dimensions[1].height = 30

    for r, fila in enumerate(filas, start=2):
        vals = [
            fila["codigo"],
            fila["archivo"],
            fila["ruta_docx"],
            fila["nombre_sistema"],
            fila["id_sistema"],
            fila["carpeta_capturas"],
            fila["npag"],
            fila["prompt"],
        ]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=(i in (2,3,4,6,8)))
            c.border = bord
            if i == 8:
                c.font = Font(size=9, name="Consolas")

    # Anchos de columna
    widths = [16, 55, 90, 55, 40, 90, 8, 100]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    for r in range(2, len(filas) + 2):
        ws.row_dimensions[r].height = 220  # espacio para el prompt

    ws.freeze_panes = "A2"

    wb.save(XLSX_OUT)
    print(f"\nExcel generado: {XLSX_OUT}")
    print(f"Total consentimientos procesados: {len(filas)}")


if __name__ == "__main__":
    main()
